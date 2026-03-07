import os
import secrets
import sqlite3
from collections import Counter
from contextlib import asynccontextmanager
from datetime import datetime
from io import BytesIO
from urllib.parse import urlencode

from dotenv import load_dotenv
from fastapi import FastAPI, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from starlette.middleware.sessions import SessionMiddleware

load_dotenv()

SECRET_KEY = os.getenv("SECRET_KEY", "change-this-secret-in-production-please")
DB_PATH = os.getenv("DB_PATH", "appointments.db")
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin123")
ADMIN_SESSION_KEY = "is_admin"

PROCEDURE_LABELS_RU = {
    "consultation": "Бесплатная консультация",
    "veneers": "Виниры E-max",
    "whitening": "Отбеливание",
    "braces": "Брекет-система",
    "implant": "Имплантация",
    "hygiene": "Профессиональная гигиена",
    "crown": "Циркониевая коронка",
    "treatment": "Лечение кариеса",
    "extraction": "Удаление зуба",
    "other": "Другое",
}


# ── Database ──────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS appointments (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            phone       TEXT    NOT NULL,
            name        TEXT    NOT NULL,
            procedure   TEXT    NOT NULL,
            created_at  TEXT    NOT NULL
        )
    """)
    conn.commit()
    conn.close()


# ── Helpers ───────────────────────────────────────────────────────────────────

def is_admin_authenticated(request: Request) -> bool:
    return request.session.get(ADMIN_SESSION_KEY) is True


def normalize_procedure_name(raw_value: str) -> str:
    value = raw_value.strip()
    if not value:
        return value

    ru_value = PROCEDURE_LABELS_RU.get(value.casefold())
    if ru_value:
        return ru_value

    for localized_name in PROCEDURE_LABELS_RU.values():
        if value.casefold() == localized_name.casefold():
            return localized_name

    return value


def normalize_date_input(raw_value: str | None, field_name: str) -> str | None:
    if raw_value is None:
        return None

    value = raw_value.strip()
    if not value:
        return None

    try:
        datetime.strptime(value, "%Y-%m-%d")
    except ValueError as exc:
        raise HTTPException(
            status_code=422,
            detail=f"Invalid {field_name}. Use YYYY-MM-DD format.",
        ) from exc

    return value


def validate_date_filters(date_from_raw: str | None, date_to_raw: str | None) -> tuple[str | None, str | None]:
    date_from = normalize_date_input(date_from_raw, "date_from")
    date_to = normalize_date_input(date_to_raw, "date_to")

    if date_from and date_to and date_from > date_to:
        raise HTTPException(status_code=422, detail="date_from must be earlier than or equal to date_to")

    return date_from, date_to


def build_appointments_query(date_from: str | None, date_to: str | None) -> tuple[str, list[str]]:
    query = "SELECT id, phone, name, procedure, created_at FROM appointments"
    clauses: list[str] = []
    params: list[str] = []

    if date_from:
        clauses.append("substr(created_at, 1, 10) >= ?")
        params.append(date_from)
    if date_to:
        clauses.append("substr(created_at, 1, 10) <= ?")
        params.append(date_to)

    if clauses:
        query += " WHERE " + " AND ".join(clauses)

    query += " ORDER BY id DESC"
    return query, params


def fetch_appointments(date_from: str | None = None, date_to: str | None = None) -> list[dict]:
    query, params = build_appointments_query(date_from, date_to)
    conn = get_db()
    try:
        rows = conn.execute(query, params).fetchall()
        items = [dict(r) for r in rows]
        for item in items:
            item["procedure"] = normalize_procedure_name(str(item["procedure"]))
        return items
    finally:
        conn.close()


# ── App ───────────────────────────────────────────────────────────────────────

@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    yield


app = FastAPI(title="Smile Avenue Dental Clinic", lifespan=lifespan)

app.add_middleware(
    SessionMiddleware,
    secret_key=SECRET_KEY,
    session_cookie="sa_session",
    max_age=3600,
    https_only=False,
    same_site="lax",
)

app.mount("/static", StaticFiles(directory="static"), name="static")

templates = Jinja2Templates(directory="templates")


# ── Public Routes ─────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    csrf_token = secrets.token_hex(32)
    request.session["csrf_token"] = csrf_token
    return templates.TemplateResponse(
        "index.html", {"request": request, "csrf_token": csrf_token}
    )


@app.post("/submit")
async def submit(
    request: Request,
    phone: str = Form(...),
    name: str = Form(...),
    procedure: str = Form(...),
    csrf_token: str = Form(...),
):
    session_token = request.session.get("csrf_token")
    if not session_token or not secrets.compare_digest(session_token, csrf_token):
        raise HTTPException(status_code=403, detail="Noto'g'ri CSRF token")

    request.session["csrf_token"] = secrets.token_hex(32)

    phone = phone.strip()
    name = name.strip()
    procedure = normalize_procedure_name(procedure)

    if not phone or not name or not procedure:
        raise HTTPException(status_code=422, detail="Barcha maydonlarni to'ldiring")

    if len(name) > 100 or len(phone) > 20 or len(procedure) > 100:
        raise HTTPException(status_code=422, detail="Kiritilgan ma'lumot juda uzun")

    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO appointments (phone, name, procedure, created_at) VALUES (?, ?, ?, ?)",
            (phone, name, procedure, datetime.now().isoformat(timespec="seconds")),
        )
        conn.commit()
    finally:
        conn.close()

    return JSONResponse({"success": True, "message": "So'rovingiz qabul qilindi!"})


# ── Admin Auth ────────────────────────────────────────────────────────────────

@app.get("/admin/login", response_class=HTMLResponse)
async def admin_login_page(request: Request):
    if is_admin_authenticated(request):
        return RedirectResponse(url="/admin", status_code=303)

    return templates.TemplateResponse(
        "admin_login.html",
        {"request": request, "error": None},
    )


@app.post("/admin/login", response_class=HTMLResponse)
async def admin_login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
):
    username = username.strip()
    password = password.strip()

    valid_username = secrets.compare_digest(username, ADMIN_USERNAME)
    valid_password = secrets.compare_digest(password, ADMIN_PASSWORD)

    if not (valid_username and valid_password):
        return templates.TemplateResponse(
            "admin_login.html",
            {
                "request": request,
                "error": "Неверный логин или пароль",
            },
            status_code=401,
        )

    request.session[ADMIN_SESSION_KEY] = True
    request.session["admin_user"] = username

    return RedirectResponse(url="/admin", status_code=303)


@app.get("/admin/logout")
async def admin_logout(request: Request):
    request.session.pop(ADMIN_SESSION_KEY, None)
    request.session.pop("admin_user", None)
    return RedirectResponse(url="/admin/login", status_code=303)


# ── Admin Data ────────────────────────────────────────────────────────────────

@app.get("/admin/appointments")
async def list_appointments(
    request: Request,
    date_from: str | None = None,
    date_to: str | None = None,
):
    if not is_admin_authenticated(request):
        raise HTTPException(status_code=401, detail="Unauthorized")

    date_from, date_to = validate_date_filters(date_from, date_to)
    rows = fetch_appointments(date_from, date_to)
    return JSONResponse(rows)


@app.get("/admin", response_class=HTMLResponse)
async def admin_dashboard(
    request: Request,
    date_from: str | None = None,
    date_to: str | None = None,
):
    if not is_admin_authenticated(request):
        return RedirectResponse(url="/admin/login", status_code=303)

    date_from, date_to = validate_date_filters(date_from, date_to)
    appointments = fetch_appointments(date_from, date_to)

    procedure_counter = Counter()
    today_prefix = datetime.now().date().isoformat()
    today_count = 0

    for item in appointments:
        procedure_counter[item["procedure"]] += 1
        if item["created_at"].startswith(today_prefix):
            today_count += 1

        try:
            item["created_at_display"] = datetime.fromisoformat(item["created_at"]).strftime("%d.%m.%Y %H:%M")
        except ValueError:
            item["created_at_display"] = item["created_at"]

    top_procedure = procedure_counter.most_common(1)[0][0] if procedure_counter else "—"

    export_query = urlencode({
        key: value
        for key, value in {
            "date_from": date_from,
            "date_to": date_to,
        }.items()
        if value
    })
    export_url = "/admin/export"
    json_url = "/admin/appointments"
    if export_query:
        export_url = f"{export_url}?{export_query}"
        json_url = f"{json_url}?{export_query}"

    return templates.TemplateResponse(
        "admin.html",
        {
            "request": request,
            "appointments": appointments,
            "total_count": len(appointments),
            "today_count": today_count,
            "top_procedure": top_procedure,
            "date_from": date_from or "",
            "date_to": date_to or "",
            "export_url": export_url,
            "json_url": json_url,
            "admin_user": request.session.get("admin_user", "admin"),
        },
    )


@app.get("/admin/export")
async def export_appointments(
    request: Request,
    date_from: str | None = None,
    date_to: str | None = None,
):
    if not is_admin_authenticated(request):
        return RedirectResponse(url="/admin/login", status_code=303)

    date_from, date_to = validate_date_filters(date_from, date_to)
    rows = fetch_appointments(date_from, date_to)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Appointments"

    headers = ["ID", "Date", "Name", "Phone", "Procedure"]
    sheet.append(headers)

    header_fill = PatternFill(start_color="C9941A", end_color="C9941A", fill_type="solid")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = header_fill

    for item in rows:
        try:
            date_text = datetime.fromisoformat(item["created_at"]).strftime("%Y-%m-%d %H:%M")
        except ValueError:
            date_text = item["created_at"]

        sheet.append([
            item["id"],
            date_text,
            item["name"],
            item["phone"],
            item["procedure"],
        ])

    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            cell_text = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(cell_text))
        sheet.column_dimensions[column].width = min(max(max_length + 2, 12), 45)

    content = BytesIO()
    workbook.save(content)

    filename = "appointments.xlsx"
    if date_from or date_to:
        left = date_from or "start"
        right = date_to or "end"
        filename = f"appointments_{left}_to_{right}.xlsx"

    return Response(
        content=content.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


# ── Run ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn

    uvicorn.run("main:app", host="0.0.0.0", port=8888, reload=True)
